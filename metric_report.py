# @title Confussion Matrix Reporter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
import pycm


def preprocess_metric(metric):
  # convert decimal to percentage with rounded 1 decimal comma
  result = []
  for key, val in metric.items():
    try:
      result.append(round(val * 100, 1))
    except:
      result.append(val)
  return result


class ExcelHandler:
  def __init__(self):
    self.wb = Workbook()
    self.ws = self.wb.active
    self.offset = 0
    self.center_alignment = Alignment(horizontal="center", vertical="center")
    self.thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

  def append(self, rekaman, lead, patologi, cm: pycm.ConfusionMatrix):
    data_entry = {
        "rekaman": rekaman,
        "lead": lead,
        "patologi": patologi,
        "overall": round(cm.Overall_ACC * 100, 1),
        "PPV": preprocess_metric(cm.PPV),
        "TPR": preprocess_metric(cm.TPR),
        "F1": preprocess_metric(cm.F1),
        "TNR": preprocess_metric(cm.TNR),
        "NPV": preprocess_metric(cm.NPV),
        "ACC": preprocess_metric(cm.ACC)
    }

    i = (self.offset * 5) + 1
    j = (self.offset * 5) + 2
    k = (self.offset * 5) + 3
    l = (self.offset * 5) + 4
    m = (self.offset * 5) + 5

    self.ws[f'A{i}'] = data_entry["rekaman"]
    self.ws[f'B{i}'] = data_entry["lead"]
    self.ws[f'C{i}'] = data_entry["patologi"]
    self.ws.merge_cells(f'A{i}:A{m}')
    self.ws.merge_cells(f'B{i}:B{m}')
    self.ws.merge_cells(f'C{i}:C{m}')

    # Kelas
    self.ws[f'D{i}'] = 'BL'
    self.ws[f'D{j}'] = 'QRS'
    self.ws[f'D{k}'] = 'T'
    self.ws[f'D{l}'] = 'P'

    # Overall Accuracy
    self.ws[f'D{m}'] = data_entry["overall"]
    self.ws.merge_cells(f'D{m}:I{m}')

    # Precision
    self.ws[f'E{i}'] = data_entry["PPV"][0]
    self.ws[f'E{j}'] = data_entry["PPV"][1]
    self.ws[f'E{k}'] = data_entry["PPV"][2]
    self.ws[f'E{l}'] = data_entry["PPV"][3]

    # Recall
    self.ws[f'F{i}'] = data_entry["TPR"][0]
    self.ws[f'F{j}'] = data_entry["TPR"][1]
    self.ws[f'F{k}'] = data_entry["TPR"][2]
    self.ws[f'F{l}'] = data_entry["TPR"][3]

    # F1-Score
    self.ws[f'G{i}'] = data_entry["F1"][0]
    self.ws[f'G{j}'] = data_entry["F1"][1]
    self.ws[f'G{k}'] = data_entry["F1"][2]
    self.ws[f'G{l}'] = data_entry["F1"][3]

    # Sensitivity
    self.ws[f'H{i}'] = data_entry["TNR"][0]
    self.ws[f'H{j}'] = data_entry["TNR"][1]
    self.ws[f'H{k}'] = data_entry["TNR"][2]
    self.ws[f'H{l}'] = data_entry["TNR"][3]

    # NPV
    self.ws[f'I{i}'] = data_entry["NPV"][0]
    self.ws[f'I{j}'] = data_entry["NPV"][1]
    self.ws[f'I{k}'] = data_entry["NPV"][2]
    self.ws[f'I{l}'] = data_entry["NPV"][3]

    # Accuracy
    self.ws[f'J{i}'] = data_entry["ACC"][0]
    self.ws[f'J{j}'] = data_entry["ACC"][1]
    self.ws[f'J{k}'] = data_entry["ACC"][2]
    self.ws[f'J{l}'] = data_entry["ACC"][3]
    self.ws[f'J{m}'] = 'Overall'

    # Apply alignment and border
    for row in self.ws.iter_rows(min_row=i, max_row=m):
      for cell in row:
        cell.alignment = self.center_alignment
        cell.border = self.thin_border

    # Increment offset for the next entry
    self.offset += 1

  def save(self, filename):
    self.wb.save(filename)
